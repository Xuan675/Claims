from __future__ import annotations

from pathlib import Path
from difflib import SequenceMatcher
from datetime import datetime, date
from typing import Optional
from openpyxl import load_workbook
import pandas as pd
import re

from payroll import (
    extract_claims_eom as eom_extract,
    reconcile_eom_line_then_total as eom_reconcile,
)
_PRIORITY = {"OK": 4, "DESC_MISMATCH": 3, "AMOUNT_MISMATCH": 3, "FLAG": 1, "NO_MATCH": 0}
LOW_INFO_TOKENS = {
    "csh","cash","reimb","reimbursement","misc","other","others",
    "fee","fees","charge","charges","voucher","claim","expense",
    "sgd","usd","myr","eur","gbp"
}
COPY_SUFFIX_RE    = re.compile(r"\(\s*\d+\s*\)$")
VERSION_SUFFIX_RE = re.compile(r"(?i)(?:[ _-]v\d+)$")
COPY_ORIG_RE = re.compile(r'(?i)\b(original|compressed|copy|scanned)\b')
DOC_PATTERNS = [
    # explicit tags first
    (re.compile(r'(?i)\bPO\s*0*([0-9]{6,})\b'),          "po_digits"),
    (re.compile(r'(?i)\bOrder\s*0*([0-9]{6,})\b'),       "order_digits"),
    (re.compile(r'(?i)\bSO\s*0*([0-9]{6,})\b'),          "so_digits"),
    (re.compile(r'(?i)\bINV(?:OICE)?\s*0*([0-9]{5,})\b'),"inv_digits"),
    (re.compile(r'(?i)\bSV0*([0-9]{6,})\b'),             "sv_digits"),
]
TRANSPORT_KEYWORDS = [
    "grab", "gojek", "tada", "uber", "taxi", "cab",
    "corporate rides", "rides", "transport", "comfort"
]

def _pick_idx_by_amount(cand_df: pd.DataFrame, target_amt: float, tol: float) -> int:
    tmp = cand_df.copy()
    tmp["_adiff"] = (tmp["Amount"] - target_amt).abs()
    within = tmp[tmp["_adiff"] <= tol]
    best = (within if not within.empty else tmp).sort_values("_adiff")
    return best.index[0]

def _route_label_for_row(common_dbg_root: Path, debug_label: str | None, pdf_desc: str) -> str:
    """
    Return a short 'FROM -> TO' label for transport rows, else ''.
    Priority:
      1) extract from debug_pages text (more reliable)
      2) fallback to the one-line pdf description
    """
    if debug_label:
        frm, to = _extract_route_from_debug_label(common_dbg_root, debug_label)
        if frm and to:
            return f" {frm} -> {to}"

    frm, to = _extract_route(pdf_desc)
    if frm and to:
        return f" {frm} -> {to}"

    return ""

def reconcile_claims_and_write_outputs(
    df: pd.DataFrame,
    args,
    claim_xlsx: str,
    support_excels: list[str],
    payroll_df: Optional[pd.DataFrame],
    common_dbg_root: Path,
    out_prefix: Path,
    start_dt: Optional[pd.Timestamp],
    end_dt: Optional[pd.Timestamp],
) -> None:
    
    out = out_prefix

    book = load_claim_book(claim_xlsx)

    for col in ("Description_pdf", "comments"):
        if col in book.columns:
            book.drop(columns=[col], inplace=True)

    original_cols = [c for c in book.columns if not c.startswith("_")]

    # normalize Amount and row index
    book["Amount"] = pd.to_numeric(book["Amount"], errors="coerce")
    if "_row_excel" not in book.columns:
        header_row0 = int(book.attrs.get("header_row0", 0))
        book["_row_excel"] = (
            book.index.to_series().astype(int) + header_row0 + 2
        )

    doccol = args.doc_col
    if doccol is not None:
        book["_doc_norm"]   = normalize_id_series(book[doccol])
        book["_doc_digits"] = book["_doc_norm"].str.replace(r"\D","", regex=True)
        book["_doc_isnum"]  = book["_doc_digits"].str.fullmatch(r"\d{7,}").fillna(False)
    else:
        book["_doc_norm"]  = ""
        book["_doc_isnum"] = False
    book["Description"] = norm_desc_series(book["Description"])
    book["_ndesc"] = book["Description"]

    matches = []
    for docno, gdoc in df.groupby("docno", sort=False):
        pool, doc_filter_used, doc_trace = _build_doc_pool(book, doccol, docno)
        if pool.empty:
            stem = str(next((x for x in gdoc["file"].unique()), ""))  # any filename in this group
            alt = extract_doc_id(Path(stem).stem, mode="first_long_digits")  # or "last_long_digits"
            if alt and alt != docno:
                pool2, why2, trace2 = _build_doc_pool(book, doccol, alt)
                if not pool2.empty:
                    docno = alt
                    pool, doc_filter_used, doc_trace = pool2, why2, trace2

        used_idx: set[int] = set()

        if args.require_doc_match and doc_filter_used == "none":
            pool = book.iloc[0:0]

        pool = pool.copy()
        gdoc = gdoc.copy()
        gdoc["_ndesc"] = gdoc["description"].map(_norm_desc_text)

        for _, row in gdoc.iterrows():
            amount = float(row["amount"])
            ndesc  = row["_ndesc"]

            if "net_zero_pair" in gdoc.columns and bool(row.get("net_zero_pair", False)):
                matches.append({
                    **row.drop(labels=["_ndesc"]).to_dict(),
                    "actual_row": None,
                    "match_desc": "To be ignored",
                    "match_amount": 0.0,
                    "match_score": 100.0,
                    "matched_tokens": "",
                    "status": "OK",
                    "doc_filter_used": "auto",
                    "desc_rule": "net_zero_pair",
                    "no_match_type": "",
                    "no_match_reason": "",
                })
                continue
            
            tol_abs = args.price_tol_abs if args.price_tol_abs is not None else args.price_tol
            tol_pct = float(args.price_tol_pct or 0.0)

            # candidates not yet used (NO fallback – enforces one-time use)
            cand = pool.copy()
            if used_idx:
                cand = cand.drop(index=[i for i in used_idx if i in cand.index], errors="ignore")
            if cand.empty:
                # Nothing unused left in the filtered pool.
                # First, try a gross-up match against the *entire* pool (since we may have filtered too hard).
                pick_idx = None
                if not pool.empty:
                    gross = round(amount * float(args.grossup), 2)
                    tol_abs_eff = (args.price_tol_abs if args.price_tol_abs is not None
                                else max(args.price_tol, abs(gross) * (args.price_tol_pct or 0.0) / 100.0))
                    cand_gross = pool[pool["Amount"].notna() &
                                    (pool["Amount"] - gross).abs() <= tol_abs_eff]
                    if not cand_gross.empty:
                        pick_idx = (cand_gross.assign(_adiff=(cand_gross["Amount"] - gross).abs())
                                            .sort_values("_adiff")
                                            .index[0])

                if pick_idx is not None:
                    # Gross-up worked → mark as OK and continue.
                    b_desc = str(book.loc[pick_idx, "Description"])
                    b_amt  = float(book.loc[pick_idx, "Amount"]) if pd.notna(book.loc[pick_idx, "Amount"]) else None
                    _, matched_tokens = desc_coverage(row["description"], b_desc)
                    best_i, best_score = best_coverage_match(row["description"], pd.Series([b_desc], index=[pick_idx]))
                    row_excel = int(book.loc[pick_idx, "_row_excel"])
                    matches.append({
                        **row.drop(labels=["_ndesc"]).to_dict(),
                        "actual_row": row_excel,
                        "match_desc": b_desc,
                        "match_amount": b_amt,
                        "match_score": round(best_score, 1),
                        "matched_tokens": " ".join(sorted(matched_tokens)),
                        "status": "OK",
                        "doc_filter_used": doc_filter_used,
                        "desc_rule": "price_grossup",
                        "no_match_type": "",
                        "no_match_reason": "",
                    })
                    used_idx.add(pick_idx)
                    continue

                # Gross-up failed too → classify the no-candidate situation.
                if pool.empty:
                    # No rows at all for this docno.
                    matches.append({
                        **row.drop(labels=["_ndesc"]).to_dict(),
                        "actual_row": None,
                        "match_desc": "", "match_amount": None,
                        "match_score": 0.0, "matched_tokens": "",
                        "status": "NO_MATCH",
                        "doc_filter_used": doc_filter_used,
                        "desc_rule": "price_required",
                        "no_match_type": "MISSING_SUPPORTING_DOCS",
                        "no_match_reason": "no Excel rows matched this docno (or doc filter produced empty pool)",
                    })
                else:
                    # There were rows, but all were already consumed.
                    matches.append({
                        **row.drop(labels=["_ndesc"]).to_dict(),
                        "actual_row": None,
                        "match_desc": "", "match_amount": None,
                        "match_score": 0.0, "matched_tokens": "",
                        "status": "NO_MATCH",
                        "doc_filter_used": doc_filter_used,
                        "desc_rule": "price_required",
                        "no_match_type": "DOC_ROWS_EXHAUSTED",
                        "no_match_reason": "all matching Excel rows for this doc were already used",
                    })
                continue

            # ---------- STRICT PRICE-FIRST ONLY ----------
            cand_amt_ok = cand[cand["Amount"].notna() & cand["Amount"].apply(
                lambda x: _within_price_tol(x, amount, tol_abs, tol_pct)    
            )]

            if cand_amt_ok.empty:
                # --- DOC-TOTAL SUM FALLBACK (for PDF totals) ---
                tol_abs_eff = (args.price_tol_abs if args.price_tol_abs is not None
                            else max(args.price_tol, abs(amount) * (args.price_tol_pct or 0.0) / 100.0))
                if len(pool) >= 2:
                    book_sum = float(pool["Amount"].fillna(0).sum())
                    if abs(book_sum - amount) <= tol_abs_eff:
                        matches.append({
                            **row.drop(labels=["_ndesc"]).to_dict(),
                            "actual_row": None,
                            "match_desc": f"Sum of all rows for {docno}",
                            "match_amount": round(book_sum, 2),
                            "match_score": 100.0,
                            "matched_tokens": "",
                            "status": "OK",
                            "doc_filter_used": doc_filter_used,
                            "desc_rule": "price_doc_total_sum",
                            "group_rows": "ALL_ROWS_IN_DOC",
                            "group_size": int(len(pool)),
                            "no_match_type": "",
                            "no_match_reason": "",
                        })
                        continue

                # --- Gross-up fallback when candidates exist but none match within tol ---
                if not pool.empty:
                    gross = round(amount * float(args.grossup), 2)
                    tol_abs_eff = (args.price_tol_abs if args.price_tol_abs is not None
                                else max(args.price_tol, abs(gross) * (args.price_tol_pct or 0.0) / 100.0))
                    cand_gross = pool[pool["Amount"].notna() &
                                    (pool["Amount"] - gross).abs() <= tol_abs_eff]
                    if not cand_gross.empty:
                        pick_idx = (cand_gross.assign(_adiff=(cand_gross["Amount"] - gross).abs())
                                            .sort_values("_adiff")
                                            .index[0])
                        b_desc = str(book.loc[pick_idx, "Description"])
                        b_amt  = float(book.loc[pick_idx, "Amount"]) if pd.notna(book.loc[pick_idx, "Amount"]) else None
                        _, matched_tokens = desc_coverage(row["description"], b_desc)
                        best_i, best_score = best_coverage_match(row["description"], pd.Series([b_desc], index=[pick_idx]))
                        row_excel = int(book.loc[pick_idx, "_row_excel"])
                        matches.append({
                            **row.drop(labels=["_ndesc"]).to_dict(),
                            "actual_row": row_excel,
                            "match_desc": b_desc,
                            "match_amount": b_amt,
                            "match_score": round(best_score, 1),
                            "matched_tokens": " ".join(sorted(matched_tokens)),
                            "status": "OK",
                            "doc_filter_used": doc_filter_used,
                            "desc_rule": "price_grossup",
                            "no_match_type": "",
                            "no_match_reason": "",
                        })
                        used_idx.add(pick_idx)
                        continue

                # Otherwise, nothing tallies.
                matches.append({
                    **row.drop(labels=["_ndesc"]).to_dict(),
                    "actual_row": None,
                    "match_desc": "", "match_amount": None,
                    "match_score": 0.0, "matched_tokens": "",
                    "status": "NO_MATCH",
                    "doc_filter_used": doc_filter_used,
                    "desc_rule": "price_required",
                    "no_match_type": "AMOUNT_DOES_NOT_TALLY",
                    "no_match_reason": "no Excel row within tolerance",
                })
                continue

            # Prefer exact normalized description within price-matched rows
            same_desc = cand_amt_ok[cand_amt_ok["_ndesc"] == ndesc]
            if not same_desc.empty:
                book_idx   = _pick_idx_by_amount(same_desc, amount, tol_abs)
                best_score = 100.0
                desc_rule  = "price_then_exact"
            else:
                # (Optional) keep only rows with >0 token overlap to avoid weak picks
                def _cov_only(d):
                    cov, _ = desc_coverage(row["description"], str(d))
                    return cov
                tmp = cand_amt_ok.copy()
                tmp["_cov"] = tmp["Description"].map(_cov_only)
                strong = tmp[tmp["_cov"] > 0]  # require any token overlap
                base = strong if not strong.empty else tmp

                # tie-break by best coverage, then by closest amount
                best_i, best_score = best_coverage_match(row["description"], base["Description"])
                book_idx = best_i if best_i in base.index else base.index[0]
                desc_rule = "price_then_coverage"

            # ---------- Scoring & status ----------
            b_desc = str(book.loc[book_idx, "Description"])
            b_amt  = float(book.loc[book_idx, "Amount"]) if pd.notna(book.loc[book_idx, "Amount"]) else None

            contains_ok = _contains_same_words(row["description"], b_desc)
            desc_ok = (best_score >= args.match_threshold) or contains_ok

            adiff = None
            relpct = None
            if b_amt is not None:
                adiff = abs(b_amt - amount)
                relpct = (adiff / abs(amount) * 100.0) if amount else None

            price_ok = _within_price_tol(b_amt, amount, tol_abs, tol_pct)

            if price_ok:
                status = "OK"
                if not desc_ok:
                    desc_rule = (desc_rule + "|weak_desc").strip("|")
            else:
                status = "AMOUNT_MISMATCH" if desc_ok else "FLAG"

            _, matched_tokens = desc_coverage(row["description"], b_desc)
            row_excel = int(book.loc[book_idx, "_row_excel"])
            matches.append({
                **row.drop(labels=["_ndesc"]).to_dict(),
                "actual_row": row_excel,
                "match_desc": b_desc,
                "match_amount": b_amt,
                "match_score": round(best_score, 1),
                "matched_tokens": " ".join(sorted(matched_tokens)),
                "status": status,
                "doc_filter_used": doc_filter_used,
                "desc_rule": desc_rule,
                "no_match_type": "",
                "no_match_reason": "",
            })

            used_idx.add(book_idx)

    md = pd.DataFrame(matches)
    # md has 'date' from the PDF item row; tag out-of-period
    if not md.empty:
        md["date"] = pd.to_datetime(md["date"], errors="coerce")
        if start_dt is not None or end_dt is not None:
            md["in_period"] = md["date"].map(lambda d: _in_period(d, start_dt, end_dt))
        else:
            md["in_period"] = True
    # --- drop MISMATCH/FLAG rows when an OK exists for same doc+price ---
    md = md.copy()

    # robust key: docno + currency + amount (2dp) + total/non-total
    md["amount_2dp"] = md["amount"].round(2)
    md["is_total"]   = md["source"].str.contains("total", case=False, na=False)
    md["_ndesc_pdf"] = md["description"].map(_norm_desc_text)
    key = ["docno", "currency", "amount_2dp", "is_total", "actual_row"]

    # status priority: higher is better
    _status_prio = {"OK": 3, "AMOUNT_MISMATCH": 2, "DESC_MISMATCH": 2, "FLAG": 1, "NO_MATCH": 0}
    md["_prio"] = md["status"].map(_status_prio).fillna(0)

    # Choose the best candidate per key by: status → match_score → (optional) description length
    md["_desc_len"] = md["description"].astype(str).str.len()

    md = (md.sort_values(key + ["_prio", "match_score", "_desc_len"],
                        ascending=[True, True, True, True, True, False, False, False])
            .drop_duplicates(subset=key, keep="first")
            .drop(columns=["_prio", "_desc_len"]))
    
    # now drop helper cols
    md.drop(columns=["amount_2dp", "is_total"], inplace=True)

    check_path = out.with_suffix(".check.xlsx")
    md.to_excel(check_path, index=False)
    print(f"Wrote: {check_path}")

    # --- choose a preferred PDF file per doc-col (soft constraint) ---
    docval_to_best_file: dict[str, str] = {}

    if doccol is not None and doccol in book.columns and not md.empty:
        tmp = md[md["actual_row"].notna()].copy()
        if not tmp.empty:
            tmp["actual_row"] = tmp["actual_row"].astype(int)

            # bring Source Doc Ref (or whatever doccol is) into md
            excel_key = book[["_row_excel", doccol]].copy()
            excel_key[doccol] = excel_key[doccol].astype(str).str.strip()

            tmp = tmp.merge(
                excel_key,
                left_on="actual_row",
                right_on="_row_excel",
                how="left",
            )

            # normalised doc-col value
            tmp["docval"] = (
                tmp[doccol]
                .astype(str)
                .str.strip()
                .replace({"nan": ""})
            )
            tmp = tmp[tmp["docval"] != ""]

            if not tmp.empty:
                # prefer files that give more OKs, then more rows total
                tmp["is_ok"] = tmp["status"].eq("OK").astype(int)

                grp = (
                    tmp.groupby(["docval", "file"], dropna=False)
                    .agg(
                        ok_count=("is_ok", "sum"),
                        total=("file", "size"),
                        avg_score=("match_score", "mean"),
                    )
                    .reset_index()
                )

                grp = grp.sort_values(
                    ["docval", "ok_count", "total", "avg_score"],
                    ascending=[True, False, False, False],
                )

                best_rows = grp.drop_duplicates(subset=["docval"], keep="first")
                docval_to_best_file = dict(
                    zip(best_rows["docval"], best_rows["file"])
                )

    # --- comments back to the claim book ---
    book_out = book.copy()

    comments    = pd.Series("NO_MATCH", index=book_out.index)
    best_status = pd.Series("NO_MATCH", index=book_out.index)
    best_rule   = pd.Series("", index=book_out.index)
    best_score  = pd.Series(pd.NA, index=book_out.index)
    best_amount = pd.Series(pd.NA, index=book_out.index)

    desc_pdf_col= pd.Series("", index=book_out.index, dtype="string")

    # (A00) DOC-LEVEL TOTAL CHECK: if Excel sum for a doc matches PDF sum, mark all its rows OK
    if doccol is not None:
        # total from PDF items (already extracted into df)
        pdf_totals = (
            df.groupby("docno", dropna=False)["amount"]
            .sum()
            .round(2)
        )

        tol_abs = args.price_tol_abs if args.price_tol_abs is not None else args.price_tol
        tol_pct = float(args.price_tol_pct or 0.0)

        for d, pdf_sum in pdf_totals.items():
            # get the Excel pool for this doc
            pool_d, _why, _trace = _build_doc_pool(book_out, doccol, d)
            if pool_d.empty:
                continue

            # Excel total for that doc
            book_sum = float(pool_d["Amount"].fillna(0).sum())
            tol_eff = max(tol_abs, abs(pdf_sum) * tol_pct / 100.0)

            if abs(book_sum - pdf_sum) <= tol_eff:
                idxs = pool_d.index
                # Do not override OUT_OF_PERIOD rows
                ok_idxs = [i for i in idxs if best_status.loc[i] != "OUT_OF_PERIOD"]

                comments.loc[ok_idxs]    = (
                    f"OK – doc total matches PDF (weak description) "
                    f"(pdf={pdf_sum:.2f}, excel={book_sum:.2f})"
                )
                best_status.loc[ok_idxs] = "OK"
                best_rule.loc[ok_idxs]   = "doc_total_match_pdf_weak_desc"
                best_score.loc[ok_idxs]  = 100.0
                best_amount.loc[ok_idxs] = book_sum

    # --- (A0) Mark net-zero groups within the claims book as OK (no PDFs needed) ---
    book_tmp = book_out.copy()
    # normalize descriptions to group robustly
    book_tmp["_ndesc"] = norm_desc_series(book_tmp["Description"])

    if doccol is not None and doccol in book_tmp.columns:
        # treat blanks/NaN as missing; we will skip those groups
        book_tmp["_dockey"] = (
            book_tmp[doccol].astype(str).str.strip().replace({"nan": ""})
        )
    else:
        book_tmp["_dockey"] = ""  # force skip

    tol_abs_cli = args.price_tol_abs if args.price_tol_abs is not None else args.price_tol
    tol_pct_cli = float(args.price_tol_pct or 0.0)

    def _eff_tol(amount_sum: float) -> float:
        # combine absolute and percentage tolerances
        return max(tol_abs_cli, abs(amount_sum) * tol_pct_cli / 100.0)
    
    # decide which date column to use in claims workbook
    claims_date_col = getattr(args, "claims_date_col", None)

    if claims_date_col and claims_date_col not in book.columns:
        print(f"[WARN] claims-date-col '{claims_date_col}' not found; falling back.")
        claims_date_col = None

    if not claims_date_col:
        for cand in ("Posting date", "Posting Date", "Document Date", "Document date"):
            if cand in book.columns:
                claims_date_col = cand
                print(f"[INFO] Using '{claims_date_col}' as claims date column.")
                break
    
    try:
        reconcile_transport_period_totals(
            book_out=book_out,
            comments=comments,
            best_status=best_status,
            best_rule=best_rule,
            best_score=best_score,
            support_excels=[Path(p) for p in support_excels],
            tol_abs=tol_abs_cli,
            tol_pct=tol_pct_cli,
            date_col=claims_date_col,
            out_prefix=args.out,
        )
    except Exception as e:
        print(f"[warn] Transport period-total reconciliation skipped: {e}")

    for (_, _), g in book_tmp.groupby(["_dockey", "_ndesc"], dropna=False):
        dockey = g["_dockey"].iloc[0]
        if not dockey:
            continue

        if len(g.index) < 2:
            continue

        amt = pd.to_numeric(g["Amount"], errors="coerce").fillna(0.0)

        # must have both signs to count as an offset
        if not ((amt > 0).any() and (amt < 0).any()):
            continue

        total = float(amt.sum())
        if abs(total) <= _eff_tol(total):
            for idx in g.index:
                # don't override OUT_OF_PERIOD
                if best_status.loc[idx] != "OUT_OF_PERIOD":
                    comments.loc[idx]    = "OK – net zero within claims book (offsetting entries)"
                    best_status.loc[idx] = "OK"
                    best_rule.loc[idx]   = "net_zero_book"
                    best_score.loc[idx]  = 100.0
                    best_amount.loc[idx] = 0.0
            continue
        
        idx_list = list(g.index)
        matched: set[int] = set()

        for i, idx_i in enumerate(idx_list):
            if idx_i in matched:
                continue
            a_i = float(amt.loc[idx_i])
            if a_i == 0.0:
                continue

            # search for a partner that cancels with idx_i
            for idx_j in idx_list[i + 1:]:
                if idx_j in matched:
                    continue
                a_j = float(amt.loc[idx_j])
                if a_j == 0.0:
                    continue

                # need opposite signs
                if a_i * a_j >= 0:
                    continue

                pair_sum = a_i + a_j
                if abs(pair_sum) <= _eff_tol(pair_sum):
                    # mark this pair as net-zero offsets
                    for idx in (idx_i, idx_j):
                        if best_status.loc[idx] != "OUT_OF_PERIOD":
                            comments.loc[idx]    = "OK – net zero within claims book (offsetting entries)"
                            best_status.loc[idx] = "OK"
                            best_rule.loc[idx]   = "net_zero_book"
                            best_score.loc[idx]  = 100.0
                            best_amount.loc[idx] = 0.0
                    matched.update({idx_i, idx_j})
                    break

    if not md.empty:
        md_used = md[md["actual_row"].notna()].copy()
        if not md_used.empty:
            md_used["actual_row"] = md_used["actual_row"].astype(int)

            status_prio = {
                "OK": 3,
                "AMOUNT_MISMATCH": 2,
                "DESC_MISMATCH": 2,
                "FLAG": 1,
                "NO_MATCH": 0,
            }
            md_used["_prio"] = md_used["status"].map(status_prio).fillna(0)

            # for each Excel row, keep the single best md record:
            # higher priority, then higher match_score
            row_best = (
                md_used.sort_values(
                    ["actual_row", "_prio", "match_score"],
                    ascending=[True, False, False],
                )
                .drop_duplicates(subset=["actual_row"], keep="first")
            )

            tol_abs_cli = (
                args.price_tol_abs if args.price_tol_abs is not None else args.price_tol
            )
            tol_pct_cli = float(args.price_tol_pct or 0.0)
            
            # track which Excel rows really have PDF support
            supported_rows = set()

            for _, best in row_best.iterrows():
                actual_row = int(best["actual_row"])
                supported_rows.add(actual_row)

                # find the row in book_out with this Excel row number
                idx_candidates = book_out.index[book_out["_row_excel"] == actual_row]
                if len(idx_candidates) == 0:
                    continue  # nothing to update

                idx0 = idx_candidates[0]
                if not (0 <= idx0 < len(book_out)):
                    continue

                # doc-col value for this row (if any)
                docval = None
                if doccol is not None and doccol in book_out.columns:
                    raw_doc = book_out.at[idx0, doccol]
                    if pd.notna(raw_doc):
                        v = str(raw_doc).strip()
                        docval = v or None

                new_status = str(best["status"])
                existing_rule = str(best_rule.iloc[idx0])

                # ALWAYS let the .check match win over the default "NO_MATCH"
                if existing_rule != "net_zero_book":
                    comments.iloc[idx0]    = _comment_from_group(
                        md_used[md_used["actual_row"] == actual_row]
                    )
                    best_status.iloc[idx0] = new_status
                    best_rule.iloc[idx0]   = str(best.get("desc_rule", ""))
                    best_score.iloc[idx0]  = float(best.get("match_score", float("nan")))
                    best_amount.iloc[idx0] = best.get("match_amount", pd.NA)

                # ---- build Description_pdf, but ONLY if price tally is OK ----
                pdf_file = str(best.get("file", "")).strip()
                pdf_page = best.get("page", None)
                try:
                    excel_desc = str(book_out.at[idx0, "Description"])
                except Exception:
                    excel_desc = ""

                pdf_desc_full = str(best.get("description", "")) or ""

                file_s = (pdf_file or "").lower()
                is_transport = (
                    _is_transport_row(excel_desc)
                    or _is_transport_row(pdf_desc_full)
                    or ("transport" in file_s)
                    or any(k in file_s for k in ("gojek", "grab", "tada", "uber", "taxi", "cab"))
                )

                # Only override file for NON-transport rows (so page->debug file stays aligned)
                if (not is_transport) and docval and docval in docval_to_best_file:
                    pdf_file = docval_to_best_file[docval]

                debug_label = None
                if pdf_file and pdf_page is not None:
                    try:
                        page_i = int(pdf_page)
                        dbg_path = (
                            common_dbg_root
                            / "debug_pages"
                            / f"{Path(pdf_file).stem}_p{page_i}.txt"
                        )
                        debug_label = dbg_path.name
                    except Exception:
                        debug_label = None

                label = ""
                if debug_label:
                    label = f"[{debug_label}]"
                elif pdf_file:
                    label = Path(pdf_file).name

                dbg_s  = (debug_label or "").lower()
                file_s = (pdf_file or "").lower()
                is_transport = (
                    _is_transport_row(excel_desc)
                    or _is_transport_row(pdf_desc_full)
                    or ("transport" in dbg_s)
                    or ("transport" in file_s)
                    or any(k in dbg_s for k in ("gojek", "grab", "tada", "uber", "taxi"))
                    or any(k in file_s for k in ("gojek", "grab", "tada", "uber", "taxi"))
                )

                # append route even if price isn't OK (route extraction is independent)
                if is_transport:
                    label += _route_label_for_row(common_dbg_root, debug_label, pdf_desc_full)

                desc_pdf_col.iloc[idx0] = label

            # clean helper column
            md_used.drop(columns=["_prio"], inplace=True, errors="ignore")

    # (B) Excel rows with NO supporting PDF at all  => "NO_MATCH, MISSING SUPPORTING DOCS"
    if "supported_rows" in locals() and supported_rows:
        supported_rows_set = set(int(r) for r in supported_rows)
    else:
        supported_rows_set = set(
            md["actual_row"].dropna().astype(int).tolist()
        )

    all_sheet_rows = set(book_out["_row_excel"].astype(int).tolist())
    missing_support_rows = sorted(all_sheet_rows - supported_rows_set)

    for r in missing_support_rows:
        idx_candidates = book_out.index[book_out["_row_excel"] == r]
        for idx in idx_candidates:
            if best_status.loc[idx] == "NO_MATCH":
                comments.loc[idx] = "NO_MATCH, MISSING SUPPORTING DOCS"
                best_status.loc[idx] = "NO_MATCH"

    # 2) DOC-LEVEL OK: if any match row says OK via doc-total,
    #    mark *all* rows in that docno as OK in the comments file.
    ok_docs = md[(md["status"].eq("OK")) &
                (md["desc_rule"].eq("price_doc_total_sum"))]["docno"].dropna().unique()

    for d in ok_docs:
        # rebuild the same pool you matched against for that docno
        pool_d, _why, _trace = _build_doc_pool(book_out, doccol, d)
        if pool_d.empty:
            continue
        sum_d = float(pool_d["Amount"].fillna(0).sum())
        idxs = pool_d.index

        comments.loc[idxs]    = f"OK – doc total matches (sum={sum_d:.2f})"
        best_status.loc[idxs] = "OK"
        best_rule.loc[idxs]   = "doc_total_sum"
        best_score.loc[idxs]  = 100.0
        best_amount.loc[idxs] = sum_d

    # ---- EOM vs Payroll (line-then-total) ----
    try:
        if payroll_df is not None and not payroll_df.empty: 
            # Extract likely EOM rows from the claims book
            claims_eom = eom_extract(
                book_df=book_out,
                date_col="Document Date",   # adjust if your claims date column differs
                amount_col="Amount",        # adjust if your claims amount column differs
            )

            # Only reconcile if we actually found EOM rows
            if not claims_eom.empty:
                rec = eom_reconcile(
                    claims_eom=claims_eom,
                    payroll_df=payroll_df,
                    start_dt=start_dt,
                    end_dt=end_dt,
                    tol_abs=0.05,  # tighten/loosen if needed
                )

                print("EOM diag:")
                print(rec["diag"])

                # 1) Line-by-line: OK rows
                for idx in rec["matched_idx"]:
                    if idx in book_out.index and best_status.loc[idx] != "OUT_OF_PERIOD":
                        comments.loc[idx]    = "OK – EOM matched to payroll (line)"
                        best_status.loc[idx] = "OK"
                        best_rule.loc[idx]   = "eom_line_match"
                        best_score.loc[idx]  = 100.0

                # 2) Totals backstop
                for idx in (rec["total_ok_idx"] - rec["matched_idx"]):
                    if idx in book_out.index and best_status.loc[idx] != "OUT_OF_PERIOD":
                        comments.loc[idx]    = "OK – EOM totals match payroll"
                        best_status.loc[idx] = "OK"
                        best_rule.loc[idx]   = "eom_total_match"
                        best_score.loc[idx]  = 100.0

                # 3) If totals didn’t match, add a single diagnostic on first EOM row
                if not rec["diag"]["totals_match"].iat[0] and len(claims_eom.index) > 0:
                    idx0 = claims_eom.index[0]
                    if idx0 in book_out.index and best_status.loc[idx0] != "OUT_OF_PERIOD":
                        c_left = rec["diag"]["claims_total_left"].iat[0]
                        p_left = rec["diag"]["payroll_total_left"].iat[0]
                        comments.loc[idx0] = (
                            f"EOM total mismatch vs payroll: "
                            f"claims_left={c_left:.2f}, payroll_left={p_left:.2f}"
                        )
                        best_rule.loc[idx0] = "eom_total_diff"
            else:
                print("[INFO] No EOM rows detected in claims workbook.")
        else:
            print("[INFO] Skipping EOM reconciliation (no payroll rows).")
    except Exception as e:
        print(f"[warn] EOM reconciliation skipped: {e}")
    
    if (start_dt is not None) or (end_dt is not None):
        if not claims_date_col or claims_date_col not in book_out.columns:
            raise KeyError(
                "Expected a date column in the claims workbook for period filtering. "
                f"Tried user-specified '{getattr(args, 'claims_date_col', None)}' and "
                "common names like 'Posting date' / 'Document Date'."
            )
        for idx in range(len(book_out)):
            excel_dt = _parse_excel_doc_date(book_out.at[idx, claims_date_col])
            if not _in_period(excel_dt, start_dt, end_dt):
                comments.iloc[idx]    = "Outside of claim period"
                best_status.iloc[idx] = "OUT_OF_PERIOD"
                best_rule.iloc[idx]   = "period_check"
    
    # (rows above the real table, or any row with blank Source Doc Ref)
    if doccol is not None and doccol in book_out.columns:
        meta_mask = (
            book_out[doccol]
            .astype(str)
            .str.strip()
            .replace({"nan": ""})
            .eq("")
        )
        # these top/metadata rows should stay blank
        comments.loc[meta_mask]    = ""
        desc_pdf_col.loc[meta_mask] = ""

    book_out = book_out[original_cols].copy()
    book_out["Description_pdf"] = (
        desc_pdf_col.reindex(book_out.index).fillna("")
    )
    book_out["comments"] = (
        comments.reindex(book_out.index).fillna("")
    )

    comments_path = str(out.with_suffix(".claims_with_comments.xlsx"))
    book_out.to_excel(comments_path, index=False)
    print(f"Wrote: {comments_path}")

def load_claim_book(path: str | Path) -> pd.DataFrame:
    """
    Load a claims workbook and automatically detect the header row.
    A valid header row MUST contain:
        - a column starting with 'description'
        - a column starting with 'amount'
    The header may appear anywhere in the sheet.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)

    ext = p.suffix.lower()

    visible_rows = None  # 1-based Excel row numbers
    if ext in {".xlsx", ".xlsm", ".xls", ".xlsb"}:
        try:
            wb = load_workbook(p, data_only=True, read_only=False)
            ws = wb[wb.sheetnames[0]]  # same sheet that read_excel will default to

            hidden = {idx for idx, dim in ws.row_dimensions.items() if dim.hidden}
            max_row = ws.max_row
            visible_rows = [r for r in range(1, max_row + 1) if r not in hidden]
        except Exception as e:
            print(f"[WARN] Could not read row visibility, using all rows: {e}")
            visible_rows = None
            
    # --- Read raw with NO HEADER so we can scan manually ---
    if ext in {".xlsx", ".xlsm", ".xls", ".xlsb"}:
        raw = pd.read_excel(p, header=None, engine="openpyxl")
    elif ext == ".csv":
        try:
            raw = pd.read_csv(
                p,
                header=None,
                dtype=str,
                engine="python",
                encoding="utf-8",
            )
        except UnicodeDecodeError:
            raw = pd.read_csv(
                p,
                header=None,
                dtype=str,
                engine="python",
                encoding="cp1252",
            )
    else:
        raise ValueError(f"Unsupported file type: {ext}")
    
    # Track real Excel row number (1-based) for each raw row
    raw["_excel_row"] = raw.index + 1

    # If we have visibility info, drop hidden rows
    if visible_rows is not None:
        raw = raw[raw["_excel_row"].isin(visible_rows)].reset_index(drop=True)

    # --- Find header row ---
    def _looks_like_header_row(vals: list[str]) -> bool:
        vals = [str(v).strip().lower() for v in vals]
        # ignore all-blank / all "nan" rows
        if all(v in ("", "nan", "none") for v in vals):
            return False

        # description-ish: "description", "short text", "text", "doc. text"
        has_desc = any(
            ("description" in v) or
            ("short text" in v) or
            (v == "text") or
            ("document header text" in v)
            for v in vals
        )
        # amount-ish: "amount", "amount in", "amt"
        has_amt = any(
            ("amount" in v) or
            (v.startswith("amt "))
            for v in vals
        )
        return has_desc and has_amt
    
    header_row = None
    for i, row in raw.drop(columns=["_excel_row"]).iterrows():
        vals = row.astype(str).tolist()
        if _looks_like_header_row(vals):
            header_row = i
            break

    if header_row is None:
        raise ValueError(
            f"Could not find header row containing 'Description' and 'Amount' in file: {p}"
        )

    # --- Rebuild dataframe using the detected header ---
    header = raw.drop(columns=["_excel_row"]).iloc[header_row].astype(str).str.strip().tolist()
    data = raw.iloc[header_row + 1:].copy()
    data.reset_index(drop=True, inplace=True)

    # excel row numbers for each data row
    excel_rows = data["_excel_row"].astype(int).to_list()
    data.drop(columns=["_excel_row"], inplace=True)

    data.columns = header
    data["_row_excel"] = excel_rows  # keep real excel row number per record

    # --- Normalize column names ---
    data.columns = [str(c).strip() for c in data.columns]

    # Normalize Amount
    if "Amount" not in data.columns:
        amt_candidates = [c for c in data.columns if str(c).lower().startswith("amount")]
        if amt_candidates:
            data.rename(columns={amt_candidates[0]: "Amount"}, inplace=True)

    # Normalize Description
    if "Description" not in data.columns:
        # 1) Exact 'description' first
        desc_candidates = [
            c for c in data.columns if str(c).strip().lower() == "description"
        ]
        # 2) Fallback: anything containing 'description'
        if not desc_candidates:
            desc_candidates = [
                c for c in data.columns if "description" in str(c).lower()
            ]
        if desc_candidates:
            data.rename(columns={desc_candidates[0]: "Description"}, inplace=True)

    # Final check
    if not {"Description", "Amount"} <= set(data.columns):    
        raise ValueError(
            f"Required columns missing after normalization. Columns found: {list(data.columns)}"
        )

    data.attrs["header_row0"] = int(header_row)
    return data

def normalize_id_series(s: pd.Series) -> pd.Series:
    def norm_one(x):
        if pd.isna(x):
            return ""
        if isinstance(x, int):
            raw = str(x)
        elif isinstance(x, float):
            raw = str(int(round(x)))
        else:
            raw = str(x).strip()
            if re.fullmatch(r"\d+(?:\.\d+)?[eE][+-]?\d+", raw):
                try:
                    raw = str(int(float(raw)))
                except Exception:
                    pass
            raw = re.sub(r"\.0+$", "", raw)
        return re.sub(r"[^A-Z0-9]", "", raw.upper())
    return s.apply(norm_one)

def _clean_place(s: str) -> str:
    s = str(s or "")

    # normalize bullets / arrows / punctuation
    s = s.replace("•", " ")
    s = re.sub(r"[\-–>]+", " ", s)
    s = re.sub(r"[\[\]\(\),;:]", " ", s)

    # remove leading time like "08:43" / "8:43AM" optionally followed by "at"
    s = re.sub(r"^\s*\d{1,2}:\d{2}\s*(?:AM|PM)?\s*(?:at\s+)?", "", s, flags=re.I)

    # remove common noise fragments found in transport receipts
    s = re.sub(r"\bdistance\s*\d+(?:\.\d+)?\s*km\b", " ", s, flags=re.I)
    s = re.sub(r"\barrived\s+on\b.*$", " ", s, flags=re.I)  # if it leaks in
    s = re.sub(r"\bpicked\s+up\s+on\b.*?\bfrom\b", " ", s, flags=re.I)

    # remove SG-style car plate (best-effort)
    s = re.sub(r"\b[A-Z]{1,3}\d{1,4}[A-Z]\b", " ", s)

    s = re.sub(r"\s+", " ", s).strip()
    return s

def _trim_to_addressish(s: str) -> str:
    """
    If string contains an address-like number, keep from the first number onward.
    Useful for Gojek where "plate • vehicle ..." precedes the pickup address.
    """
    if not s:
        return s
    m = re.search(r"\b\d{1,4}[A-Za-z]?\b", s)
    return s[m.start():].strip() if m else s.strip()


def _dedupe_tail(s: str) -> str:
    """
    Gojek sometimes repeats the pickup address:
      "... 26B Jalan Membina ... 26B Jalan Membina"
    Keep the last occurrence.
    """
    if not s:
        return s
    # if the last 3+ tokens appear earlier, prefer the last span
    toks = s.split()
    if len(toks) < 6:
        return s
    tail = " ".join(toks[-3:])
    pos = s.lower().rfind(tail.lower())
    if pos > 0:
        # keep from a bit earlier so we don't cut mid-word
        return s[pos:].strip()
    return s

def _strip_provider_prefix(s: str) -> str:
    s = str(s).strip()
    if not s:
        return s
    parts = s.split(None, 1)
    if parts and parts[0].lower().rstrip(":") in TRANSPORT_KEYWORDS:
        return parts[1] if len(parts) > 1 else ""
    return s

def _extract_route(text: str) -> tuple[str, str]:
    """
    Try to extract 'from' and 'to' places from a transport description string.
    Returns (from_place, to_place); empty strings if not detected.
    """
    if not text:
        return "", ""

    s = _strip_provider_prefix(str(text))

    # 1) "from X to Y"
    m = re.search(r"(?i)\bfrom\s+(.+?)\s+to\s+(.+)", s)
    if m:
        return _clean_place(m.group(1)), _clean_place(m.group(2))

    # 2) "to Y from X"  (swap)
    m = re.search(r"(?i)\bto\s+(.+?)\s+from\s+(.+)", s)
    if m:
        return _clean_place(m.group(2)), _clean_place(m.group(1))

    # 3) "X to Y" (whole string)
    m = re.search(r"(?i)^(.+?)\s+to\s+(.+)$", s)
    if m:
        return _clean_place(m.group(1)), _clean_place(m.group(2))

    # 4) "X - Y" / "X – Y" / "X > Y"
    m = re.search(r"(?i)^(.+?)\s*[-–>]\s*(.+)$", s)
    if m:
        return _clean_place(m.group(1)), _clean_place(m.group(2))

    return "", ""

def _extract_route_from_page_text(text: str) -> tuple[str, str]:
    """
    Extract 'from' and 'to' from the FULL page text.
    Handles:
      - Gojek: '... from <addr> ... Arrived on ... at <addr> ...'
      - Grab:  'Your Trip' section with origin/destination lines.
    """
    if not text:
        return "", ""

    s = re.sub(r"\s+", " ", text)

    # --- GOJEK-style receipts (robust cleanup) ---
    m = re.search(
        r"(?is)\bfrom\s+(.+?)\s+\bDistance\b.*?\b\d{1,2}:\d{2}\s*(?:AM|PM)?\s+at\s+(.+?)(?:\s+\bHelp\b|\s+\bReport\b|$)",
        s,
    )
    if m:
        raw_from = m.group(1)
        raw_to   = m.group(2)
        frm = _dedupe_tail(_trim_to_addressish(_clean_place(raw_from))).strip()
        to  = _clean_place(raw_to).strip()
        if frm and to:
            return frm, to

    # --- Grab-style: keep your existing logic below (optionally improve is_time) ---
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    start_idx = None
    for i, ln in enumerate(lines):
        if re.search(r"(?i)\byour\s+trip\b", ln):
            start_idx = i
            break

    def is_time(ln: str) -> bool:
        # allow leading symbols like "⋮ 8:48AM"
        ln2 = re.sub(r"^[^\d]+", "", ln).strip()
        return bool(re.match(r"^\d{1,2}:\d{2}\s*(?:AM|PM)?$", ln2, flags=re.I))

    if start_idx is not None:
        stops: list[str] = []
        for i in range(start_idx + 1, len(lines)):
            ln = lines[i]

            if re.search(r"(?i)follow us|©\s*grab|got an issue|help centre|fare breakdown", ln):
                break

            if is_time(ln):
                prev = lines[i - 1].strip() if i - 1 > start_idx else ""
                nxt  = lines[i + 1].strip() if i + 1 < len(lines) else ""

                cand = prev
                if (not cand) or re.fullmatch(r"[⋮•.]+", cand):
                    cand = nxt

                # filter noise
                if cand and not is_time(cand):
                    if not re.search(
                        r"(?i)lost item|incorrect trip|incorrect fare|get support|your trip|"
                        r"\d+\s*km\b|\bmins?\b",
                        cand,
                    ):
                        stops.append(cand)

            if len(stops) >= 2:
                break

        if len(stops) >= 2:
            return _clean_place(stops[0]), _clean_place(stops[1])

    # generic fallback
    m = re.search(r"(?is)\bfrom\s+(.+?)\s+\bto\s+(.+?)(?:\s{2,}|$)", s)
    if m:
        return _clean_place(m.group(1)), _clean_place(m.group(2))

    return "", ""

def _extract_route_from_debug_label(debug_root: Path, debug_label: str) -> tuple[str, str]:
    """
    Load debug_pages/<debug_label> and extract a route, if any.
    """
    def _read(lbl: str) -> str:
        p = debug_root / "debug_pages" / lbl
        return p.read_text(encoding="utf-8", errors="ignore")
    
    # 1) current page
    try:
        text = _read(debug_label)
        frm, to = _extract_route_from_page_text(text)
        if frm and to:
            return frm, to
    except Exception:
        pass

    m = re.search(r"_p(\d+)\.txt$", debug_label)
    if not m:
        return "", ""

    n = int(m.group(1))
    for k in (n - 1, n, n + 1, n + 2):
        if k < 0:
            continue
        alt = re.sub(r"_p\d+\.txt$", f"_p{k}.txt", debug_label)
        if alt == debug_label:
            continue
        try:
            text = _read(alt)
            frm, to = _extract_route_from_page_text(text)
            if frm and to:
                return frm, to
        except Exception:
            continue

    return "", ""

def _norm_desc_text(text: str) -> str:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    s = str(text)

    s = s.lower()
    s = re.sub(r"[-/]", " ", s)           # replace dash/slash with space
    s = re.sub(r"[^a-z0-9 ]+", " ", s)    # keep only a-z, 0-9 and spaces
    s = re.sub(r"\s+", " ", s).strip()    # collapse spaces
    return s

def norm_desc_series(s: pd.Series) -> pd.Series:
    return s.map(_norm_desc_text).fillna("")

def desc_coverage(extracted_desc: str, excel_desc: str) -> tuple[float, set[str]]:
    nx = _norm_desc_text(extracted_desc)
    ne = _norm_desc_text(excel_desc)
    if not ne:
        return 1.0, set()
    if ne in nx:
        etoks = _desc_tokens(excel_desc)
        return 1.0, set(etoks)

    etoks = _desc_tokens(excel_desc)
    xtoks = set(_desc_tokens(extracted_desc))
    if not etoks:
        return 1.0, set()

    matched = []
    total_w = 0
    hit_w = 0
    for t in etoks:
        w = 2 if any(c.isdigit() for c in t) else 1
        total_w += w
        if t in xtoks:
            hit_w += w
            matched.append(t)
    return (hit_w / max(1, total_w)), set(matched)

def _desc_tokens(text: str) -> list[str]:
    return [t for t in _norm_desc_text(text).split() if t and t not in STOPWORDS]

STOPWORDS = {
    "the","a","an","and","or","of","for","to","in","on","at","by",
    "with","from","this","that","these","those",
    "invoice","receipt","qty","quantity","pcs","unit","item","items"
}

def best_coverage_match(extracted_desc: str, pool: pd.Series) -> tuple[int, float]:
    best_i, best_cov = -1, 0.0
    for i, txt in pool.items():
        cov, _ = desc_coverage(extracted_desc, str(txt))
        if cov > best_cov:
            best_cov = cov
            best_i = i
    return best_i, best_cov * 100.0

def _contains_same_words(pdf_desc: str, excel_desc: str) -> bool:
    """
    True if all meaningful Excel tokens appear in the PDF tokens (order-free).
    Allows simple plural/singular variants. Treat empty/low-info Excel desc as OK.
    """
    pdf_set = set(_desc_tokens(pdf_desc))
    # keep only meaningful words from the Excel description
    sig = [w for w in _desc_tokens(excel_desc) if w not in LOW_INFO_TOKENS]
    if not sig:   # Excel desc has nothing but low-info words -> treat as OK
        return True

    def variants(w: str) -> set[str]:
        v = {w}
        if w.endswith("es"): v.add(w[:-2])
        if w.endswith("s"):  v.add(w[:-1])
        return v

    for w in sig:
        if not any(v in pdf_set for v in variants(w)):
            return False
    return True

def _within_price_tol(book_amt: Optional[float], pdf_amt: float,
                      tol_abs: float, tol_pct: float) -> bool:
    if book_amt is None or pd.isna(book_amt):
        return False
    adiff = abs(float(book_amt) - float(pdf_amt))
    pct_ok = (tol_pct > 0 and pdf_amt and (adiff / abs(pdf_amt) * 100.0) <= tol_pct)
    return (adiff <= tol_abs) or pct_ok

def build_doc_keys(file_stem: str) -> tuple[set[str], set[str]]:
    base = norm_docno(file_stem or "")
    eq: set[str] = set()
    contains: set[str] = set()
    if not base:
        return eq, contains
    
    eq.add(base)
    eq.add(base.lstrip("0"))

    for sub in re.findall(r"[A-Z]{1,4}\d{3,}", base):
        eq.add(sub)
        eq.add(sub.lstrip("0"))
        
    tokens = re.findall(r"[A-Z0-9]{4,}", base)
    for t in tokens:
        # full token can be an exact key
        eq.add(t)
        eq.add(t.lstrip("0"))

        for num in re.findall(r"\d{5,}", t):
            contains.add(num)
            contains.add(num.lstrip("0"))
            if len(num) >= 8:
                contains.add(num[-8:])
                contains.add(num[-9:])

    # 2) pure numeric substrings (for long invoice IDs etc.)
    for num in re.findall(r"\d{5,}", base):
        eq.add(num)
        eq.add(num.lstrip("0"))

    return eq, contains

def norm_docno(s: str) -> str:
    """Normalize a document id from a filename-like string (drop copy/version suffixes)."""
    s = clean_file_stem(str(s))
    return re.sub(r"[^A-Z0-9]", "", s.upper())

def clean_file_stem(stem: str) -> str:
    s = stem.strip()
    s = COPY_SUFFIX_RE.sub("", s)
    s = VERSION_SUFFIX_RE.sub("", s)
    s = COPY_ORIG_RE.sub("", s)
    return s.strip(" _-")

def _build_doc_pool(book: pd.DataFrame, doccol: Optional[str], pdf_docno: str) -> tuple[pd.DataFrame, str, str]:
    if not doccol or not pdf_docno:
        return book, "none", "no-doccol-or-docno"

    pdf_norm = norm_docno(pdf_docno or "")
    eq_tokens, contains_tokens = build_doc_keys(pdf_norm)

    mask_eq = book["_doc_norm"].isin(eq_tokens)
    if mask_eq.any():
        pool = book[mask_eq]
        return pool, "doc_eq", f"eq={mask_eq.sum()}"

    scores = book["_doc_norm"].map(lambda x: _doc_suffix_score(pdf_norm, x))
    strong = (scores >= 8) & book["_doc_isnum"]
    if strong.any():
        best = scores[strong].max() 
        pool = book[(scores == best) & strong]
        return pool, f"doc_suffix_{int(best)}", f"suffix_max={int(best)} kept={len(pool)}"

    mask_ct = pd.Series(False, index=book.index)
    for t in contains_tokens:
        mask_ct |= (book["_doc_isnum"] &
                    book["_doc_norm"].str.contains(t, na=False))
    if mask_ct.any():
        pool = book[mask_ct]
        return pool, "doc_contains", f"contains kept={mask_ct.sum()} tokens={list(contains_tokens)}"

    alpha_tokens = re.findall(r"[A-Z]{2,}\d{3,}", pdf_norm)
    if alpha_tokens:
        mask_alpha = book["_doc_norm"].isin(alpha_tokens)
        if mask_alpha.any():
            pool = book[mask_alpha]
            return pool, "doc_alpha", f"alpha={alpha_tokens} kept={mask_alpha.sum()}"

    overlap_scores = book["_doc_norm"].map(
        lambda x: _doc_overlap_fraction(pdf_norm, x)
    )

    best = overlap_scores.max()
    THRESH = 0.8  # or 0.9 if you want stricter

    if best >= THRESH:
        mask_frac = overlap_scores >= THRESH
        pool = book[mask_frac]
        return (
            pool,
            f"doc_frac_{best:.2f}",
            f"overlap_fraction>={THRESH} kept={mask_frac.sum()}",
        )

    # 6) final fallback – no doc filter
    return book, "none", "no-doc-match"

def _doc_suffix_score(pdf_docno: str, excel_doc: str) -> int: 
    a = re.sub(r"\D", "", norm_docno(pdf_docno or "")) 
    b = re.sub(r"\D", "", str(excel_doc or "")) 
    i = 0 
    while i < min(len(a), len(b)) and a[-1 - i] == b[-1 - i]: 
        i += 1
    return i

def _doc_overlap_fraction(pdf_docno: str, excel_doc: str) -> float:
    """
    Return the fraction of the *shorter* normalized ID covered by the
    longest common substring **that contains at least one digit**.
    
    0.0 means 'no meaningful overlap'.
    1.0 means 'shorter ID is fully contained in the longer one'.
    """
    a = norm_docno(pdf_docno or "")
    b = norm_docno(str(excel_doc or ""))

    if not a or not b:
        return 0.0

    m = SequenceMatcher(None, a, b).find_longest_match(0, len(a), 0, len(b))
    if m.size == 0:
        return 0.0

    common = a[m.a : m.a + m.size]

    # require at least one digit to avoid matching purely alpha stuff like "INV"
    if not any(ch.isdigit() for ch in common):
        return 0.0

    shorter_len = min(len(a), len(b))
    return m.size / shorter_len

def _comment_from_group(grp: pd.DataFrame) -> str:
    g = grp.copy()
    g["__prio"] = g["status"].map(_PRIORITY).fillna(0)
    best = g.sort_values(["__prio", "match_score"], ascending=[False, False]).iloc[0]

    st = str(best["status"])
    via = str(best.get("doc_filter_used", "none"))
    scr = best.get("match_score", None)
    a   = float(best.get("amount", 0.0))
    b   = best.get("match_amount", None)
    diff = (abs(a - float(b)) if (b is not None and pd.notna(b)) else None)
    row_excel = best.get("actual_row", None)
    row_hint = f" (row {int(row_excel)})" if pd.notna(row_excel) else ""

    desc_rules = g.get("desc_rule", "").astype(str)
    is_grossup = desc_rules.str.lower().eq("price_grossup").any()
    doc_total  = desc_rules.str.contains("doc_total", case=False, na=False).any() \
                 or desc_rules.str.contains("doc_total_match_pdf", case=False, na=False).any()

    if st == "OK" or doc_total:
        if is_grossup:
            return f"OK (GROSSUP MATCH){row_hint}"
        k = int((g["status"] == "OK").sum())
        return f"OK – {k} match(es) via {via}{row_hint}"
    if st == "DESC_MISMATCH":
        return f"OK – price ok, weak description (score {scr}){row_hint}"
    if st == "AMOUNT_MISMATCH":
        return f"FLAG – description ok (score {scr}), price diff {diff:.2f}{row_hint}"
    if st == "FLAG":
        return f"FLAG – low match (score {scr}){row_hint}"
    return "NO_MATCH"

def _parse_excel_doc_date(val):
    if isinstance(val, (pd.Timestamp, datetime, date)):
        return pd.to_datetime(val, errors="coerce")
    # Excel serials (numbers or digit-strings)
    try:
        if isinstance(val, (int, float)) and not pd.isna(val):
            # Excel epoch (Windows): 1899-12-30
            return pd.to_datetime("1899-12-30") + pd.to_timedelta(int(val), unit="D")
        s = str(val).strip()
        if s.isdigit():
            return pd.to_datetime("1899-12-30") + pd.to_timedelta(int(s), unit="D")
    except Exception:
        pass
    # text formats (assume dayfirst for 31/07/2025)
    s = str(val).strip()
    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        for fmt in ("%d/%m/%Y","%d-%m-%Y","%Y-%m-%d","%d %b %Y","%d %B %Y","%d/%m/%y","%d-%m-%y"):
            try:
                from datetime import datetime as _dt
                return pd.to_datetime(_dt.strptime(s, fmt).date())
            except Exception:
                continue
    return dt  # may be NaT

def _in_period(dt, start_dt, end_dt):
    if pd.isna(dt):
        return True
    if start_dt is not None and dt < start_dt:
        return False
    if end_dt   is not None and dt > end_dt:
        return False
    return True

def reconcile_transport_period_totals(
    book_out,
    comments,
    best_status,
    best_rule,
    best_score,
    support_excels,
    tol_abs,
    tol_pct,
    date_col=None,
    out_prefix=None,
):
    """Mark transport rows OK when GL period total matches a support Excel total."""
    print(f"[DBG] transport_recon: out_prefix={out_prefix}")
    print(f"[DBG] transport_recon: support_excels={len(support_excels)} files")

    if not support_excels or not date_col or date_col not in book_out.columns:
        print("[DBG] transport_recon: early return – "
            f"support_excels={bool(support_excels)}, "
            f"has_DocDate={'Document Date' in book_out.columns}")
        return
    
    summaries = _summarise_support_excels(support_excels)
    print(f"[DBG] transport_recon: summaries_built={len(summaries)}")
    if not summaries:
        print("[DBG] transport_recon: early return – no summaries from support_excels")
        return
    
    wbs_col = next(
        (c for c in ["WBS Element", "WBS", "WBS Element Code"] if c in book_out.columns),
        None,
    )

    tmp = book_out.copy()
    tmp[date_col] = pd.to_datetime(tmp[date_col], errors="coerce")
    tmp["period_key"] = tmp[date_col].dt.to_period("M")
    tmp["_is_transport"] = tmp["Description"].apply(_is_transport_row)

    gl_transport = tmp[tmp["_is_transport"] & tmp[date_col].notna()]
    if gl_transport.empty:
        return

    for period, g in gl_transport.groupby("period_key"):
        amt = pd.to_numeric(g["Amount"], errors="coerce").fillna(0.0)
        gl_total = float(amt.sum().round(2))
        tol_eff = max(tol_abs, abs(gl_total) * tol_pct / 100.0)

        gl_wbs = set()
        if wbs_col:
            vals = g[wbs_col].dropna().astype(str).str.strip()
            gl_wbs.update(vals[vals != ""].tolist())

        # find first support summary with matching period and close total
        best = None
        best_overlap = -1
        for s in summaries:
            if s["period"] != period:
                continue
            if abs(s["total"] - gl_total) > tol_eff:
                continue

            overlap = len(gl_wbs & s["wbs_set"]) if gl_wbs and s["wbs_set"] else 0
            if overlap > best_overlap:
                best_overlap = overlap
                best = s

        if not best:
            continue

        for idx in g.index:
            if str(best_status.iloc[idx]) == "OUT_OF_PERIOD":
                continue
            comments.iloc[idx] = (
                f"OK – period transport total matches support "
                f"{best['file'].name} [{best['sheet']}] "
                f"(period={period}, gl={gl_total:.2f}, support={best['total']:.2f})"
            )
            best_status.iloc[idx] = "OK"
            best_rule.iloc[idx] = "transport_period_total"
            best_score.iloc[idx] = 100.0

def extract_doc_id(
    stem: str,
    mode: str = "auto",
    segment_sep: str = r"[ _-]",
    segment_index: int = 0,
) -> str:
    s = clean_file_stem(stem)

    # 1) explicit modes
    if mode == "po":
        m = re.search(r'(?i)\bPO\s*0*([0-9]{6,})\b', s);  return m.group(1) if m else ""
    if mode == "order":
        m = re.search(r'(?i)\bOrder\s*0*([0-9]{6,})\b', s); return m.group(1) if m else ""
    if mode in ("invoice","inv"):
        m = re.search(r'(?i)\bINV(?:OICE)?\s*0*([0-9]{5,})\b', s); return m.group(1) if m else ""
    if mode == "sv":
        m = re.search(r'(?i)\bSV0*([0-9]{6,})\b', s); return ("SV" + m.group(1)) if m else ""

    if mode.startswith("regex:"):
        # e.g. mode="regex:(?i)_(SV\d{6,})_:1" or "regex:(\d{8,}):1"
        try:
            pat, grp = mode.split(":", 2)[1:]
            grp = int(grp)
            m = re.search(pat, s)
            return m.group(grp) if m else ""
        except Exception:
            return ""

    if mode == "segment":
        parts = re.split(segment_sep, s)
        return parts[segment_index] if 0 <= segment_index < len(parts) else ""

    if mode in ("first_long_digits","last_long_digits"):
        runs = re.findall(r'(\d{8,})', s)
        if runs:
            return runs[0] if mode == "first_long_digits" else runs[-1]
        return ""

    # 2) auto: try tagged first
    for rx, _name in DOC_PATTERNS:
        m = rx.search(s)
        if m:
            # include tag if it’s alnum prefixed like SVxxxx, else just digits
            if _name.startswith("sv_"):
                return "SV" + m.group(1)
            return m.group(1)

    # 3) longest digit run (>=8). If multiple, pick the longest; tie -> rightmost
    runs = re.findall(r'(\d{8,})', s)
    if runs:
        max_len = max(len(r) for r in runs)
        candidates = [r for r in runs if len(r) == max_len]
        return candidates[-1]

    # 4) alnum tag like SV01954455 / INV123456 if no pure digit runs
    m = re.search(r'([A-Z]{1,4}\d{6,})', s, flags=re.I)
    if m:
        return m.group(1).upper()

    # 5) last resort: your old normalized stem
    return norm_docno(s)

def _is_transport_row(desc) -> bool:
    d = str(desc).lower()
    return any(k in d for k in TRANSPORT_KEYWORDS)

def _summarise_support_excels(support_paths):
    """Return list of {period, total, wbs_set, file, sheet} from support workbooks."""
    month_rx = re.compile(r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s*([0-9]{2,4})",
                          re.I)
    mon_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
               "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

    out = []
    for path in support_paths:
        try:
            xls = pd.ExcelFile(path)
        except Exception:
            continue

        for sheet in xls.sheet_names:
            try:
                df = xls.parse(sheet)
            except Exception:
                continue
            if df.empty:
                continue

            # period from any date column, else from filename
            period = None
            for c in df.columns:
                if "date" in str(c).lower():
                    dt = pd.to_datetime(df[c], errors="coerce").dropna()
                    if not dt.empty:
                        period = dt.iloc[0].to_period("M")
                        break
            if period is None:
                m = month_rx.search(path.name)
                if not m:
                    continue
                mon_s, year_s = m.groups()
                month = mon_map[mon_s.lower()]
                y = int(year_s)
                year = 2000 + y if y < 100 else y
                period = pd.Period(year=year, month=month, freq="M")

            total_candidates = [c for c in df.columns
                                if "total" in str(c).lower() or "amount" in str(c).lower()]

            if total_candidates:
                # prefer the *right-most* "total" column
                # (works for WBS / Sum of Fare / 2% Admin / Total layouts)
                total_col = total_candidates[-1]
            else:
                # 2) fallback: numeric column with the biggest absolute sum
                num_cols = []
                for c in df.columns:
                    s = pd.to_numeric(df[c], errors="coerce")
                    if s.notna().any():
                        num_cols.append((c, float(s.fillna(0).sum())))
                if not num_cols:
                    continue
                # choose column whose sum has largest absolute value
                total_col = max(num_cols, key=lambda t: abs(t[1]))[0]

            total_series = pd.to_numeric(df[total_col], errors="coerce")
            if total_series.notna().sum() == 0:
                continue
            total = float(total_series.fillna(0).sum().round(2))

            wbs_set = set()
            for c in df.columns:
                if "wbs" in str(c).lower():
                    vals = df[c].dropna().astype(str).str.strip()
                    wbs_set.update(vals[vals != ""].tolist())

            out.append(
                {
                    "period": period,
                    "total": total,
                    "wbs_set": wbs_set,
                    "file": path,
                    "sheet": sheet,
                }
            )
    return out